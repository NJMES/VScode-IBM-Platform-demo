import csv
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import config

_COLUMNS = [
    "id",
    "needs_review",
    "source",
    "review_status",
    "domain",
    "question_type",
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_answer",
    "explanation",
    "linkedin_post_url",
    "created_at",
]

_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FONT = Font(bold=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _row_to_dict(row: sqlite3.Row) -> dict:
    d = dict(row)
    d["needs_review"] = "Yes" if d.get("review_status") == config.REVIEW_STATUS_PENDING else "No"
    return d


def export_csv(
    rows: list[sqlite3.Row],
    output_dir: Path,
    filename_prefix: str = "cissp_questions",
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{filename_prefix}_{_timestamp()}.csv"

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(_row_to_dict(row))

    return path


def export_excel(
    rows: list[sqlite3.Row],
    output_dir: Path,
    filename_prefix: str = "cissp_questions",
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{filename_prefix}_{_timestamp()}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "All Questions"

    # Header row
    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=False)

    ws.freeze_panes = "A2"

    # Data rows
    correct_answer_col_idx = _COLUMNS.index("correct_answer") + 1
    question_text_col_idx = _COLUMNS.index("question_text") + 1

    for row_idx, row in enumerate(rows, start=2):
        d = _row_to_dict(row)
        is_pending = d.get("review_status") == config.REVIEW_STATUS_PENDING

        for col_idx, col_name in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=d.get(col_name))

            # Yellow fill for AI questions needing review
            if is_pending:
                cell.fill = _YELLOW

            # Green fill for the correct answer cell (validated questions only)
            if col_idx == correct_answer_col_idx and not is_pending and d.get("correct_answer"):
                cell.fill = _GREEN

            # Wrap text for question and explanation columns
            if col_name in ("question_text", "explanation"):
                cell.alignment = Alignment(wrap_text=True)

    # Auto-filter on all columns
    ws.auto_filter.ref = ws.dimensions

    # Column widths
    col_widths = {
        "id": 6, "needs_review": 13, "source": 20, "review_status": 15,
        "domain": 35, "question_type": 12, "question_text": 60,
        "option_a": 40, "option_b": 40, "option_c": 40, "option_d": 40,
        "correct_answer": 14, "explanation": 60,
        "linkedin_post_url": 30, "created_at": 20,
    }
    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    # Row height for question rows
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 60

    wb.save(path)
    return path


def export_both(
    rows: list[sqlite3.Row],
    output_dir: Optional[Path] = None,
    filename_prefix: str = "cissp_questions",
) -> tuple[Path, Path]:
    out = output_dir or config.EXPORTS_DIR
    csv_path = export_csv(rows, out, filename_prefix)
    xlsx_path = export_excel(rows, out, filename_prefix)
    return csv_path, xlsx_path
